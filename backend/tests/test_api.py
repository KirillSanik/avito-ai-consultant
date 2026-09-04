import os
from pathlib import Path
import tempfile
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import select

test_db_path = Path(tempfile.gettempdir()) / f"reviewdesk-test-{uuid4().hex}.db"
os.environ["DATABASE_URL"] = f"sqlite:///{test_db_path}"

from app.database import Base, SessionLocal, engine, ensure_schema
from app.main import app, seed_demo_data
from app.models import AuthToken, Course, Submission, User


@pytest.fixture()
def client() -> TestClient:
    Base.metadata.drop_all(engine)
    Base.metadata.create_all(engine)
    seed_demo_data()
    with TestClient(app) as test_client:
        yield test_client


def test_auth_register_login_and_me(client: TestClient) -> None:
    login = client.post(
        "/api/auth/login",
        json={"login": "reviewer", "password": "reviewer"},
    )
    assert login.status_code == 200
    assert login.json()["user"]["role"] == "reviewer"

    token = login.json()["token"]
    me = client.get(
        "/api/auth/me",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert me.status_code == 200
    assert me.json()["login"] == "reviewer"
    assert client.get("/api/auth/me").status_code == 401

    registered = client.post(
        "/api/auth/register",
        json={
            "login": "new_reviewer",
            "password": "safe-password",
            "first_name": "Новый",
            "last_name": "Ревьюер",
            "telegram": "@new_reviewer",
            "role": "reviewer",
        },
    )
    assert registered.status_code == 201
    assert registered.json()["user"]["login"] == "new_reviewer"
    assert client.post(
        "/api/auth/login",
        json={"login": "new_reviewer", "password": "wrong"},
    ).status_code == 401
    assert client.post(
        "/api/auth/register",
        json={
            "login": "new_reviewer",
            "password": "another-password",
            "first_name": "Другой",
            "last_name": "Пользователь",
            "telegram": "@other",
            "role": "methodist",
        },
    ).status_code == 409

    with SessionLocal() as db:
        user = db.scalar(select(User).where(User.login == "new_reviewer"))
        assert user is not None
        assert user.password_hash != "safe-password"
        assert user.password_hash.startswith("pbkdf2_sha256$")


def test_auth_rejects_unknown_user_and_wrong_password(client: TestClient) -> None:
    unknown = client.post(
        "/api/auth/login",
        json={"login": "ghost", "password": "reviewer"},
    )
    assert unknown.status_code == 401
    assert unknown.json()["detail"] == "Invalid login or password"

    wrong = client.post(
        "/api/auth/login",
        json={"login": "reviewer", "password": "not-the-password"},
    )
    assert wrong.status_code == 401
    assert client.get("/api/auth/me").status_code == 401
    assert client.get(
        "/api/auth/me",
        headers={"Authorization": "Bearer invalid-token"},
    ).status_code == 401


def test_auth_survives_app_restart(client: TestClient) -> None:
    registered = client.post(
        "/api/auth/register",
        json={
            "login": "persist_user",
            "password": "keep-me",
            "first_name": "Сохранённый",
            "last_name": "Аккаунт",
            "telegram": "@persist_user",
            "role": "reviewer",
        },
    )
    assert registered.status_code == 201
    token = registered.json()["token"]

    with SessionLocal() as db:
        user = db.scalar(select(User).where(User.login == "persist_user"))
        stored_tokens = db.scalars(select(AuthToken)).all()
        assert user is not None
        assert stored_tokens

    with TestClient(app) as restarted:
        still_me = restarted.get(
            "/api/auth/me",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert still_me.status_code == 200
        assert still_me.json()["login"] == "persist_user"

        login = restarted.post(
            "/api/auth/login",
            json={"login": "persist_user", "password": "keep-me"},
        )
        assert login.status_code == 200
        assert login.json()["user"]["id"] == registered.json()["user"]["id"]


def _auth_headers(client: TestClient, login: str, password: str) -> dict[str, str]:
    token = client.post(
        "/api/auth/login",
        json={"login": login, "password": password},
    ).json()["token"]
    return {"Authorization": f"Bearer {token}"}


def _review_payload(assignment: dict, scores: list[int] | None = None) -> dict:
    criteria = assignment["criteria"]
    values = scores if scores is not None else [item["max_score"] for item in criteria]
    return {
        "criterion_scores": [
            {
                "criterion_index": index,
                "score": values[index],
                "comment": f"Комментарий {index + 1}",
            }
            for index, _ in enumerate(criteria)
        ],
        "summary": "Проверено по критериям, работа принимается.",
        "integrity_flag": None,
    }


def _xlsx_bytes(*logins: str) -> bytes:
    from io import BytesIO

    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["login"])
    for login in logins:
        sheet.append([login])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_courses_and_assignments_contract(client: TestClient) -> None:
    assert client.get("/api/courses").status_code == 401
    methodist = _auth_headers(client, "methodist", "methodist")
    reviewer = _auth_headers(client, "reviewer", "reviewer")

    courses = client.get("/api/courses", headers=methodist)
    assert courses.status_code == 200
    assert len(courses.json()) == 2
    active_course = next(item for item in courses.json() if item["active"])
    assert {
        "stream",
        "active",
        "cover_color",
        "students_count",
        "assignments_count",
    } <= active_course.keys()

    reviewer_courses = client.get("/api/courses", headers=reviewer).json()
    assert len(reviewer_courses) == 1
    assert reviewer_courses[0]["id"] == active_course["id"]

    active = client.get("/api/courses", params={"active": "true"}, headers=methodist).json()
    completed = client.get("/api/courses", params={"active": "false"}, headers=methodist).json()
    assert len(active) == 1
    assert len(completed) == 1
    assert completed[0]["active"] is False
    assert client.get("/api/courses", params={"active": "false"}, headers=reviewer).json() == []

    assignments = client.get(
        f"/api/courses/{active_course['id']}/assignments",
        headers=methodist,
    )
    assert assignments.status_code == 200
    first = assignments.json()[0]
    assert {
        "number",
        "task_url",
        "criteria_url",
        "total",
        "reviewed",
        "reviewer_checked",
        "reviewer_total",
    } <= first.keys()

    assignment = client.get(
        f"/api/assignments/{first['id']}", headers=methodist
    )
    assert assignment.status_code == 200
    assert assignment.json()["number"] == first["number"]
    assert assignment.json()["criteria_url"] == first["criteria_url"]


def test_methodist_can_create_course_reviewer_cannot(client: TestClient) -> None:
    reviewer_headers = _auth_headers(client, "reviewer", "reviewer")
    forbidden = client.post(
        "/api/courses",
        headers=reviewer_headers,
        json={
            "title": "Новый поток",
            "year": 2026,
            "cohort": "Зимний поток",
            "stream": 8,
        },
    )
    assert forbidden.status_code == 403

    anonymous = client.post(
        "/api/courses",
        json={
            "title": "Новый поток",
            "year": 2026,
            "cohort": "Зимний поток",
            "stream": 8,
        },
    )
    assert anonymous.status_code == 401

    created = client.post(
        "/api/courses",
        headers=_auth_headers(client, "methodist", "methodist"),
        json={
            "title": "SQL для аналитиков",
            "year": 2026,
            "cohort": "Зимний поток",
            "stream": 8,
            "cover_color": "#0891B2",
            "students_count": 32,
        },
    )
    assert created.status_code == 201
    body = created.json()
    assert body["title"] == "SQL для аналитиков"
    assert body["stream"] == 8
    assert body["active"] is True
    assert body["assignments_count"] == 0

    listed = client.get(
        "/api/courses",
        params={"active": "true"},
        headers=_auth_headers(client, "methodist", "methodist"),
    ).json()
    assert any(item["id"] == body["id"] for item in listed)
    reviewer_listed = client.get("/api/courses", headers=reviewer_headers).json()
    assert all(item["id"] != body["id"] for item in reviewer_listed)

    with SessionLocal() as db:
        stored = db.get(Course, body["id"])
        assert stored is not None
        assert stored.title == "SQL для аналитиков"


def test_ensure_schema_adds_missing_course_columns(client: TestClient) -> None:
    from sqlalchemy import inspect, text

    inspector = inspect(engine)
    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE courses DROP COLUMN stream"))
    inspector.clear_cache()
    assert "stream" not in {column["name"] for column in inspector.get_columns("courses")}

    ensure_schema()
    inspector.clear_cache()
    assert "stream" in {column["name"] for column in inspector.get_columns("courses")}

    created = client.post(
        "/api/courses",
        headers=_auth_headers(client, "methodist", "methodist"),
        json={
            "title": "Курс после миграции колонок",
            "year": 2026,
            "cohort": "Тестовый поток",
            "stream": 3,
        },
    )
    assert created.status_code == 201
    assert created.json()["stream"] == 3


def test_main_reviewer_flow(client: TestClient) -> None:
    headers = _auth_headers(client, "reviewer", "reviewer")
    course_id = client.get(
        "/api/courses", params={"active": "true"}, headers=headers
    ).json()[0]["id"]
    assignment_id = client.get(
        f"/api/courses/{course_id}/assignments",
        headers=headers,
    ).json()[0]["id"]
    assignment = client.get(
        f"/api/assignments/{assignment_id}", headers=headers
    )
    body = assignment.json()
    submission_id = body["submissions"][0]["id"]

    draft = client.post(
        f"/api/submissions/{submission_id}/ai-draft", headers=headers
    )
    assert draft.status_code == 200
    assert draft.json()["ai_draft"]["total"] > 0

    too_high = list(item["max_score"] for item in body["criteria"])
    too_high[0] = too_high[0] + 1
    assert client.put(
        f"/api/submissions/{submission_id}/review",
        headers=headers,
        json=_review_payload(body, too_high),
    ).status_code == 422

    review = client.put(
        f"/api/submissions/{submission_id}/review",
        headers=headers,
        json=_review_payload(body, [20, 10, 5, 4]),
    )
    assert review.status_code == 200
    assert review.json()["status"] == "reviewed"
    assert review.json()["score"] == 39
    assert len(review.json()["criterion_scores"]) == 4

    report = client.get(
        f"/api/submissions/{submission_id}/report.pdf", headers=headers
    )
    assert report.status_code == 200
    assert report.headers["content-type"] == "application/pdf"


def test_reviewer_management_and_clarification_patch(client: TestClient) -> None:
    headers = _auth_headers(client, "methodist", "methodist")
    course_id = client.get(
        "/api/courses", params={"active": "true"}, headers=headers
    ).json()[0]["id"]
    assignment_id = client.get(
        f"/api/courses/{course_id}/assignments",
        headers=headers,
    ).json()[0]["id"]

    seeded_reviewers = client.get(
        f"/api/assignments/{assignment_id}/reviewers", headers=headers
    )
    assert seeded_reviewers.status_code == 200
    assert len(seeded_reviewers.json()) == 3

    extra = client.post(
        "/api/auth/register",
        json={
            "login": "hw_reviewer",
            "password": "safe-password",
            "first_name": "Мария",
            "last_name": "Соколова",
            "telegram": "@maria_hw",
            "role": "reviewer",
        },
    )
    assert extra.status_code == 201
    extra_id = extra.json()["user"]["id"]
    assert client.post(
        f"/api/assignments/{assignment_id}/reviewers",
        headers=headers,
        json={"user_id": extra_id},
    ).status_code == 404

    assert client.post(
        f"/api/courses/{course_id}/reviewers",
        headers=headers,
        json={"user_id": extra_id},
    ).status_code == 201
    created = client.post(
        f"/api/assignments/{assignment_id}/reviewers",
        headers=headers,
        json={"user_id": extra_id},
    )
    assert created.status_code == 201
    assert created.json()["name"] == "Мария Соколова"
    assert created.json()["user_id"] == extra_id
    assert client.post(
        f"/api/assignments/{assignment_id}/reviewers",
        headers=headers,
        json={"user_id": extra_id},
    ).status_code == 409
    reviewer_id = created.json()["id"]
    assert client.delete(
        f"/api/assignments/{assignment_id}/reviewers/{reviewer_id}",
        headers=headers,
    ).status_code == 204
    assert client.delete(
        f"/api/assignments/{assignment_id}/reviewers/{reviewer_id}",
        headers=headers,
    ).status_code == 404

    clarification = client.post(
        f"/api/assignments/{assignment_id}/clarifications",
        headers=_auth_headers(client, "reviewer", "reviewer"),
        json={"message": "Нужно ли учитывать альтернативный способ расчёта?"},
    )
    assert clarification.status_code == 200
    clarification_id = clarification.json()["id"]
    updated = client.patch(
        f"/api/clarifications/{clarification_id}",
        headers=headers,
        json={"status": "accepted"},
    )
    assert updated.status_code == 200
    assert updated.json()["status"] == "accepted"
    assert client.patch(
        f"/api/clarifications/{clarification_id}",
        headers=headers,
        json={"status": "open"},
    ).status_code == 422

    dashboard = client.get("/api/dashboard", headers=headers)
    assert dashboard.status_code == 200
    assert len(dashboard.json()["reviewers"]) >= 1
    assert len(dashboard.json()["clarifications"]) >= 1


def test_methodist_assigns_reviewer_to_course(client: TestClient) -> None:
    methodist = _auth_headers(client, "methodist", "methodist")
    reviewer = _auth_headers(client, "reviewer", "reviewer")

    catalog = client.get("/api/reviewers", headers=methodist)
    assert catalog.status_code == 200
    logins = {item["login"] for item in catalog.json()}
    assert "reviewer" in logins
    assert "methodist" in logins
    assert client.get("/api/reviewers", headers=reviewer).status_code == 403

    extra = client.post(
        "/api/auth/register",
        json={
            "login": "second_reviewer",
            "password": "safe-password",
            "first_name": "Второй",
            "last_name": "Ревьюер",
            "telegram": "@second_reviewer",
            "role": "reviewer",
        },
    )
    assert extra.status_code == 201
    extra_user_id = extra.json()["user"]["id"]
    extra_headers = {"Authorization": f"Bearer {extra.json()['token']}"}
    assert client.get("/api/courses", headers=extra_headers).json() == []

    created = client.post(
        "/api/courses",
        headers=methodist,
        json={
            "title": "Курс только для назначенных",
            "year": 2026,
            "cohort": "Закрытый поток",
            "stream": 4,
        },
    )
    assert created.status_code == 201
    course_id = created.json()["id"]
    assert client.get(
        f"/api/courses/{course_id}/assignments",
        headers=reviewer,
    ).status_code == 403
    assert all(
        item["id"] != course_id
        for item in client.get("/api/courses", headers=reviewer).json()
    )

    assigned = client.get(f"/api/courses/{course_id}/reviewers", headers=methodist)
    assert assigned.status_code == 200
    assert assigned.json() == []

    demo_id = next(item["id"] for item in catalog.json() if item["login"] == "reviewer")
    added = client.post(
        f"/api/courses/{course_id}/reviewers",
        headers=methodist,
        json={"user_id": demo_id},
    )
    assert added.status_code == 201
    assert added.json()["login"] == "reviewer"
    assert client.post(
        f"/api/courses/{course_id}/reviewers",
        headers=methodist,
        json={"user_id": demo_id},
    ).status_code == 409
    assert client.post(
        f"/api/courses/{course_id}/reviewers",
        headers=reviewer,
        json={"user_id": extra_user_id},
    ).status_code == 403
    assert client.post(
        f"/api/courses/{course_id}/reviewers",
        headers=methodist,
        json={"user_id": extra.json()["user"]["id"] + 1000},
    ).status_code == 404

    visible = client.get("/api/courses", headers=reviewer).json()
    assert any(item["id"] == course_id for item in visible)
    assert client.get(
        f"/api/courses/{course_id}/assignments",
        headers=reviewer,
    ).status_code == 200
    assert all(
        item["id"] != course_id
        for item in client.get("/api/courses", headers=extra_headers).json()
    )

    client.post(
        f"/api/courses/{course_id}/reviewers",
        headers=methodist,
        json={"user_id": extra_user_id},
    )
    assert any(
        item["id"] == course_id
        for item in client.get("/api/courses", headers=extra_headers).json()
    )
    assert client.delete(
        f"/api/courses/{course_id}/reviewers/{extra_user_id}",
        headers=methodist,
    ).status_code == 204
    assert all(
        item["id"] != course_id
        for item in client.get("/api/courses", headers=extra_headers).json()
    )


def test_methodist_reviewer_tab_requires_self_assignment(client: TestClient) -> None:
    methodist = _auth_headers(client, "methodist", "methodist")
    me = client.get("/api/auth/me", headers=methodist).json()
    all_courses = client.get("/api/courses", headers=methodist).json()
    assert len(all_courses) == 2
    assert client.get(
        "/api/courses",
        params={"as_reviewer": "true"},
        headers=methodist,
    ).json() == []

    course_id = next(item["id"] for item in all_courses if item["active"])
    assigned = client.post(
        f"/api/courses/{course_id}/reviewers",
        headers=methodist,
        json={"user_id": me["id"]},
    )
    assert assigned.status_code == 201
    assert assigned.json()["login"] == "methodist"

    reviewing = client.get(
        "/api/courses",
        params={"as_reviewer": "true"},
        headers=methodist,
    ).json()
    assert [item["id"] for item in reviewing] == [course_id]
    assert len(client.get("/api/courses", headers=methodist).json()) == 2

    reviewer = _auth_headers(client, "reviewer", "reviewer")
    hidden_id = next(item["id"] for item in all_courses if not item["active"])
    reviewer_courses = client.get(
        "/api/courses",
        params={"as_reviewer": "false"},
        headers=reviewer,
    ).json()
    assert all(item["id"] != hidden_id for item in reviewer_courses)


def test_methodist_can_create_homework_reviewer_cannot(client: TestClient) -> None:
    methodist = _auth_headers(client, "methodist", "methodist")
    reviewer = _auth_headers(client, "reviewer", "reviewer")
    course_id = client.get(
        "/api/courses", params={"active": "true"}, headers=methodist
    ).json()[0]["id"]
    payload = {
        "title": "Новое домашнее задание",
        "deadline": "2026-10-01T23:59:00",
        "task_url": "https://github.com/example/new-homework",
        "criteria_url": "",
        "criteria": [
            {
                "title": "Корректность расчётов",
                "max_score": 60,
                "description": "Проверьте формулы и воспроизводимость.",
            },
            {
                "title": "Выводы",
                "max_score": 40,
                "description": "Выводы должны опираться на данные.",
            },
        ],
    }
    assert client.post(
        f"/api/courses/{course_id}/assignments",
        json=payload,
    ).status_code == 401
    assert client.post(
        f"/api/courses/{course_id}/assignments",
        headers=reviewer,
        json=payload,
    ).status_code == 403

    created = client.post(
        f"/api/courses/{course_id}/assignments",
        headers=methodist,
        json=payload,
    )
    assert created.status_code == 201
    body = created.json()
    assert body["title"] == "Новое домашнее задание"
    assert body["number"] == 3
    assert body["total"] == 0

    listed = client.get(
        f"/api/courses/{course_id}/assignments",
        headers=methodist,
    ).json()
    assert any(item["id"] == body["id"] for item in listed)
    assignment = client.get(
        f"/api/assignments/{body['id']}", headers=methodist
    ).json()
    assert assignment["criteria_url"] == ""
    assert assignment["criteria"][0]["title"] == "Корректность расчётов"
    assert assignment["criteria"][0]["description"] == "Проверьте формулы и воспроизводимость."
    assert assignment["criteria"][1]["max_score"] == 40

    enrolled = client.get(f"/api/courses/{course_id}/reviewers", headers=methodist).json()
    demo_id = next(item["user_id"] for item in enrolled if item["login"] == "reviewer")
    with_reviewers = client.post(
        f"/api/courses/{course_id}/assignments",
        headers=methodist,
        json={
            **payload,
            "title": "ДЗ с ревьюерами курса",
            "reviewer_user_ids": [demo_id],
        },
    )
    assert with_reviewers.status_code == 201
    assigned = client.get(
        f"/api/assignments/{with_reviewers.json()['id']}/reviewers",
        headers=methodist,
    ).json()
    assert any(item["user_id"] == demo_id for item in assigned)

    too_much = client.post(
        f"/api/courses/{course_id}/assignments",
        headers=methodist,
        json={
            **payload,
            "title": "Слишком много баллов",
            "criteria": [
                {"title": "Первый", "max_score": 60, "description": ""},
                {"title": "Второй", "max_score": 50, "description": ""},
            ],
        },
    )
    assert too_much.status_code == 422
    assignment_id = client.get(
        f"/api/courses/{course_id}/assignments",
        headers=methodist,
    ).json()[0]["id"]
    overflow_update = client.put(
        f"/api/assignments/{assignment_id}/criteria",
        headers=methodist,
        json={
            "reviewer_guide": "Проверьте работу по критериям подробно.",
            "criteria": [
                {"title": "Первый", "max_score": 80, "description": ""},
                {"title": "Второй", "max_score": 30, "description": ""},
            ],
        },
    )
    assert overflow_update.status_code == 422


def test_criteria_must_total_exactly_100(client: TestClient) -> None:
    methodist = _auth_headers(client, "methodist", "methodist")
    reviewer = _auth_headers(client, "reviewer", "reviewer")
    course_id = client.get("/api/courses", headers=methodist).json()[0]["id"]
    assignment_id = client.get(
        f"/api/courses/{course_id}/assignments", headers=methodist
    ).json()[0]["id"]

    def update(total: int, headers: dict[str, str] | None = methodist):
        return client.put(
            f"/api/assignments/{assignment_id}/criteria",
            headers=headers,
            json={
                "reviewer_guide": "Подробная инструкция для ревьюера.",
                "criteria": [
                    {
                        "title": "Критерий один",
                        "max_score": total,
                        "description": "Описание",
                    }
                ],
            },
        )

    assert update(99).status_code == 422
    assert update(101).status_code == 422
    assert update(100).status_code == 200
    assert update(100, reviewer).status_code == 403
    assert update(100, None).status_code == 401


def test_reviewer_only_sees_assigned_homeworks_and_bulk_add(
    client: TestClient,
) -> None:
    methodist = _auth_headers(client, "methodist", "methodist")
    reviewer = _auth_headers(client, "reviewer", "reviewer")
    course_id = client.get(
        "/api/courses", params={"active": "true"}, headers=methodist
    ).json()[0]["id"]
    all_homeworks = client.get(
        f"/api/courses/{course_id}/assignments", headers=methodist
    ).json()
    reviewer_homeworks = client.get(
        f"/api/courses/{course_id}/assignments",
        params={"as_reviewer": "true"},
        headers=reviewer,
    ).json()
    assert len(all_homeworks) == 2
    assert len(reviewer_homeworks) == 1
    hidden_id = next(
        item["id"]
        for item in all_homeworks
        if item["id"] not in {visible["id"] for visible in reviewer_homeworks}
    )
    assert client.get(
        f"/api/assignments/{hidden_id}",
        params={"as_reviewer": "true"},
        headers=reviewer,
    ).status_code == 403

    reviewer_user_id = client.get("/api/auth/me", headers=reviewer).json()["id"]
    bulk = client.post(
        f"/api/assignments/{hidden_id}/reviewers/bulk",
        headers=methodist,
        json={"user_ids": [reviewer_user_id]},
    )
    assert bulk.status_code == 200
    assert bulk.json()[0]["user_id"] == reviewer_user_id
    assert client.post(
        f"/api/assignments/{hidden_id}/reviewers/bulk",
        headers=methodist,
        json={"user_ids": [reviewer_user_id]},
    ).json() == []

    now_visible = client.get(
        f"/api/courses/{course_id}/assignments", headers=reviewer
    ).json()
    assert {item["id"] for item in now_visible} == {
        item["id"] for item in all_homeworks
    }
    detail = client.get(
        f"/api/assignments/{hidden_id}", headers=reviewer
    )
    assert detail.status_code == 200
    assert detail.json()["reviewer_guide"]


def test_methodist_dismisses_reviewer_suggestion(client: TestClient) -> None:
    methodist = _auth_headers(client, "methodist", "methodist")
    reviewer = _auth_headers(client, "reviewer", "reviewer")
    course_id = client.get("/api/courses", headers=reviewer).json()[0]["id"]
    assignment_id = client.get(
        f"/api/courses/{course_id}/assignments", headers=reviewer
    ).json()[0]["id"]
    created = client.post(
        f"/api/assignments/{assignment_id}/clarifications",
        headers=reviewer,
        json={"message": "Проверьте формулировку второго критерия."},
    )
    assert created.status_code == 200
    clarification_id = created.json()["id"]
    assert client.patch(
        f"/api/clarifications/{clarification_id}",
        headers=reviewer,
        json={"status": "dismissed"},
    ).status_code == 403
    dismissed = client.patch(
        f"/api/clarifications/{clarification_id}",
        headers=methodist,
        json={"status": "dismissed"},
    )
    assert dismissed.status_code == 200
    assert dismissed.json()["status"] == "dismissed"


def test_student_portal_application_enrollment_and_submission(
    client: TestClient,
) -> None:
    methodist = _auth_headers(client, "methodist", "methodist")
    registered = client.post(
        "/api/student/auth/register",
        json={
            "login": "student_one",
            "password": "student-pass",
            "first_name": "Анна",
            "last_name": "Студентова",
            "telegram": "@student_one",
        },
    )
    assert registered.status_code == 201
    assert registered.json()["user"]["role"] == "student"
    student = {"Authorization": f"Bearer {registered.json()['token']}"}
    assert client.get("/api/student/auth/me", headers=student).status_code == 200
    assert client.post(
        "/api/student/auth/login",
        json={"login": "methodist", "password": "methodist"},
    ).status_code == 401
    assert client.post(
        "/api/auth/register",
        json={
            "login": "wrong_student_route",
            "password": "password",
            "first_name": "Wrong",
            "last_name": "Route",
            "telegram": "wrong",
            "role": "student",
        },
    ).status_code == 422

    courses = client.get("/api/student/courses", headers=student)
    assert courses.status_code == 200
    active = courses.json()[0]
    assert active["capacity"] == 30
    assert active["enrolled_count"] == 0
    assert active["enrollment_status"] == "none"
    course_id = active["id"]
    detail = client.get(
        f"/api/student/courses/{course_id}", headers=student
    ).json()
    assert detail["description"]
    assert detail["assignments"] == []
    assignment_id = client.get(
        f"/api/courses/{course_id}/assignments", headers=methodist
    ).json()[0]["id"]
    assert client.get(
        f"/api/student/assignments/{assignment_id}", headers=student
    ).status_code == 403

    applied = client.post(
        f"/api/student/courses/{course_id}/apply", headers=student
    )
    assert applied.status_code == 201
    assert applied.json()["status"] == "pending"
    assert client.post(
        f"/api/student/courses/{course_id}/apply", headers=student
    ).status_code == 409
    applications = client.get(
        "/api/enrollment-applications",
        params={"status": "pending"},
        headers=methodist,
    )
    assert applications.status_code == 200
    application_id = applications.json()[0]["id"]
    approved = client.patch(
        f"/api/enrollment-applications/{application_id}",
        headers=methodist,
        json={"status": "enrolled"},
    )
    assert approved.status_code == 200
    assert approved.json()["status"] == "enrolled"
    assert client.patch(
        f"/api/enrollment-applications/{application_id}",
        headers=methodist,
        json={"status": "rejected"},
    ).status_code == 409

    mine = client.get("/api/student/courses/mine", headers=student).json()
    assert [item["id"] for item in mine] == [course_id]
    enrolled_detail = client.get(
        f"/api/student/courses/{course_id}", headers=student
    ).json()
    assert enrolled_detail["enrollment_status"] == "enrolled"
    deadlines = [item["deadline"] for item in enrolled_detail["assignments"]]
    assert deadlines == sorted(deadlines)
    assert all("reviewer_guide" not in item for item in enrolled_detail["assignments"])

    assert client.post(
        f"/api/student/assignments/{assignment_id}/submit",
        headers=student,
        json={"work_url": "http://github.com/example/insecure"},
    ).status_code == 422
    assert client.post(
        f"/api/student/assignments/{assignment_id}/submit",
        headers=student,
        json={"work_url": "https://evil.example/work"},
    ).status_code == 422
    submitted = client.post(
        f"/api/student/assignments/{assignment_id}/submit",
        headers=student,
        json={"work_url": "https://github.com/example/student-work"},
    )
    assert submitted.status_code == 200
    assert submitted.json()["status"] == "pending"
    assert client.post(
        f"/api/student/assignments/{assignment_id}/submit",
        headers=student,
        json={"work_url": "https://github.com/example/student-work-2"},
    ).status_code == 409
    student_homework = client.get(
        f"/api/student/assignments/{assignment_id}", headers=student
    ).json()
    assert "criteria" not in student_homework
    assert "criteria_url" not in student_homework
    assert "reviewer_guide" not in student_homework

    with SessionLocal() as db:
        stored = db.scalar(
            select(Submission).where(
                Submission.assignment_id == assignment_id,
                Submission.student_user_id == registered.json()["user"]["id"],
            )
        )
        assert stored is not None
        stored.status = "reviewed"
        stored.score = 87
        db.commit()
    points = client.get(
        f"/api/student/courses/{course_id}", headers=student
    ).json()["total_points"]
    assert points == 87


def test_student_rejection_capacity_and_submission_privacy(
    client: TestClient,
) -> None:
    methodist = _auth_headers(client, "methodist", "methodist")

    def register(login: str) -> tuple[dict, int]:
        response = client.post(
            "/api/student/auth/register",
            json={
                "login": login,
                "password": "student-pass",
                "first_name": login,
                "last_name": "Student",
                "telegram": "",
            },
        )
        return (
            {"Authorization": f"Bearer {response.json()['token']}"},
            response.json()["user"]["id"],
        )

    first, _ = register("capacity_one")
    second, _ = register("capacity_two")
    course = client.post(
        "/api/courses",
        headers=methodist,
        json={
            "title": "Курс с одним местом",
            "year": 2026,
            "cohort": "Пилот",
            "stream": 9,
            "capacity": 1,
            "description": "Проверка вместимости.",
        },
    ).json()
    course_id = course["id"]
    first_application = client.post(
        f"/api/student/courses/{course_id}/apply", headers=first
    ).json()
    assert client.patch(
        f"/api/enrollment-applications/{first_application['id']}",
        headers=methodist,
        json={"status": "enrolled"},
    ).status_code == 200
    assert client.post(
        f"/api/student/courses/{course_id}/apply", headers=second
    ).status_code == 409

    rejected_course = client.post(
        "/api/courses",
        headers=methodist,
        json={
            "title": "Курс для отклонения",
            "year": 2026,
            "cohort": "Пилот",
            "stream": 10,
            "capacity": 2,
        },
    ).json()
    rejected = client.post(
        f"/api/student/courses/{rejected_course['id']}/apply", headers=second
    ).json()
    assert client.patch(
        f"/api/enrollment-applications/{rejected['id']}",
        headers=methodist,
        json={"status": "rejected"},
    ).status_code == 200
    detail = client.get(
        f"/api/student/courses/{rejected_course['id']}", headers=second
    ).json()
    assert detail["enrollment_status"] == "rejected"
    assert client.post(
        f"/api/student/courses/{rejected_course['id']}/apply", headers=second
    ).status_code == 201


def test_lifespan_does_not_seed_by_default() -> None:
    Base.metadata.drop_all(engine)
    Base.metadata.create_all(engine)
    with TestClient(app) as test_client:
        assert test_client.get("/health").status_code == 200
        assert test_client.post(
            "/api/auth/login",
            json={"login": "reviewer", "password": "reviewer"},
        ).status_code == 401


def test_course_description_acl_and_applications_filter(
    client: TestClient,
) -> None:
    methodist = _auth_headers(client, "methodist", "methodist")
    reviewer = _auth_headers(client, "reviewer", "reviewer")
    outsider = client.post(
        "/api/auth/register",
        json={
            "login": "desc_outsider",
            "password": "safe-password",
            "first_name": "Внешний",
            "last_name": "Ревьюер",
            "telegram": "@desc_outsider",
            "role": "reviewer",
        },
    )
    outsider_headers = {
        "Authorization": f"Bearer {outsider.json()['token']}",
    }
    course = client.post(
        "/api/courses",
        headers=methodist,
        json={
            "title": "Курс с описанием",
            "year": 2026,
            "cohort": "Пилот",
            "stream": 11,
            "description": "Старое описание",
        },
    ).json()
    course_id = course["id"]
    assert client.patch(
        f"/api/courses/{course_id}",
        headers=outsider_headers,
        json={"description": "Нельзя"},
    ).status_code == 403
    assert client.post(
        f"/api/courses/{course_id}/reviewers",
        headers=methodist,
        json={"user_id": client.get("/api/auth/me", headers=reviewer).json()["id"]},
    ).status_code == 201
    updated = client.patch(
        f"/api/courses/{course_id}",
        headers=reviewer,
        json={"description": "Описание от ревьюера"},
    )
    assert updated.status_code == 200
    assert updated.json()["description"] == "Описание от ревьюера"
    methodist_update = client.patch(
        f"/api/courses/{course_id}",
        headers=methodist,
        json={"description": "Описание методиста"},
    )
    assert methodist_update.status_code == 200

    student = client.post(
        "/api/student/auth/register",
        json={
            "login": "app_filter_student",
            "password": "student-pass",
            "first_name": "Фильтр",
            "last_name": "Заявок",
            "telegram": "",
        },
    )
    student_headers = {
        "Authorization": f"Bearer {student.json()['token']}",
    }
    client.post(f"/api/student/courses/{course_id}/apply", headers=student_headers)
    other = client.post(
        "/api/courses",
        headers=methodist,
        json={
            "title": "Другой курс",
            "year": 2026,
            "cohort": "Пилот",
            "stream": 12,
        },
    ).json()
    client.post(
        f"/api/student/courses/{other['id']}/apply",
        headers=student_headers,
    )
    all_pending = client.get(
        "/api/enrollment-applications",
        params={"status": "pending"},
        headers=methodist,
    ).json()
    filtered = client.get(
        "/api/enrollment-applications",
        params={"status": "pending", "course_id": course_id},
        headers=methodist,
    ).json()
    assert len(all_pending) >= 2
    assert {item["course_id"] for item in filtered} == {course_id}


def test_reviewer_history_and_submission_modal(client: TestClient) -> None:
    methodist = _auth_headers(client, "methodist", "methodist")
    reviewer = _auth_headers(client, "reviewer", "reviewer")
    course_id = client.get(
        "/api/courses", params={"active": "true"}, headers=reviewer
    ).json()[0]["id"]
    assignment_id = client.get(
        f"/api/courses/{course_id}/assignments", headers=reviewer
    ).json()[0]["id"]
    assignment = client.get(
        f"/api/assignments/{assignment_id}", headers=reviewer
    ).json()
    own_ids = {item["id"] for item in assignment["submissions"]}
    assert own_ids
    assert all(
        item["reviewer_user_id"]
        == client.get("/api/auth/me", headers=reviewer).json()["id"]
        for item in assignment["submissions"]
    )
    submission_id = assignment["submissions"][0]["id"]
    reviewed = client.put(
        f"/api/submissions/{submission_id}/review",
        headers=reviewer,
        json=_review_payload(assignment, [10, 10, 5, 2]),
    )
    assert reviewed.status_code == 200
    history = client.get(
        f"/api/assignments/{assignment_id}", headers=reviewer
    ).json()["submissions"]
    assert any(
        item["id"] == submission_id and item["status"] == "reviewed"
        for item in history
    )
    mine = client.get(
        f"/api/submissions/{submission_id}", headers=reviewer
    )
    assert mine.status_code == 200
    assert mine.json()["criterion_scores"]
    staff = client.get(
        f"/api/submissions/{submission_id}", headers=methodist
    )
    assert staff.status_code == 200

    extra = client.post(
        "/api/auth/register",
        json={
            "login": "other_queue",
            "password": "safe-password",
            "first_name": "Другая",
            "last_name": "Очередь",
            "telegram": "@other_queue",
            "role": "reviewer",
        },
    )
    extra_headers = {
        "Authorization": f"Bearer {extra.json()['token']}",
    }
    extra_id = extra.json()["user"]["id"]
    client.post(
        f"/api/courses/{course_id}/reviewers",
        headers=methodist,
        json={"user_id": extra_id},
    )
    client.post(
        f"/api/assignments/{assignment_id}/reviewers",
        headers=methodist,
        json={"user_id": extra_id},
    )
    assert client.get(
        f"/api/submissions/{submission_id}", headers=extra_headers
    ).status_code == 403
    nxt = client.get(
        f"/api/assignments/{assignment_id}/next", headers=extra_headers
    )
    if nxt.status_code == 200:
        assert nxt.json()["reviewer_user_id"] == extra_id


def test_rebalance_and_xlsx_roundtrip(client: TestClient) -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    methodist = _auth_headers(client, "methodist", "methodist")
    first = client.post(
        "/api/auth/register",
        json={
            "login": "balance_one",
            "password": "safe-password",
            "first_name": "Первый",
            "last_name": "Баланс",
            "telegram": "@balance_one",
            "role": "reviewer",
        },
    ).json()["user"]
    second = client.post(
        "/api/auth/register",
        json={
            "login": "balance_two",
            "password": "safe-password",
            "first_name": "Второй",
            "last_name": "Баланс",
            "telegram": "@balance_two",
            "role": "reviewer",
        },
    ).json()["user"]
    course = client.post(
        "/api/courses",
        headers=methodist,
        json={
            "title": "Курс баланса",
            "year": 2026,
            "cohort": "Пилот",
            "stream": 13,
            "capacity": 10,
        },
    ).json()
    course_id = course["id"]
    for user in (first, second):
        assert client.post(
            f"/api/courses/{course_id}/reviewers",
            headers=methodist,
            json={"user_id": user["id"]},
        ).status_code == 201
    homework = client.post(
        f"/api/courses/{course_id}/assignments",
        headers=methodist,
        json={
            "title": "Баланс проверки",
            "deadline": "2026-10-01T23:59:00",
            "task_url": "https://github.com/example/task",
            "reviewer_guide": "Проверьте работу по критериям внимательно.",
            "criteria": [
                {"title": "Первый", "max_score": 60, "description": ""},
                {"title": "Второй", "max_score": 40, "description": ""},
            ],
            "reviewer_user_ids": [first["id"]],
        },
    ).json()
    assignment_id = homework["id"]

    def enroll_and_submit(login: str) -> None:
        registered = client.post(
            "/api/student/auth/register",
            json={
                "login": login,
                "password": "student-pass",
                "first_name": login,
                "last_name": "Student",
                "telegram": "",
            },
        )
        headers = {
            "Authorization": f"Bearer {registered.json()['token']}",
        }
        application = client.post(
            f"/api/student/courses/{course_id}/apply", headers=headers
        ).json()
        client.patch(
            f"/api/enrollment-applications/{application['id']}",
            headers=methodist,
            json={"status": "enrolled"},
        )
        submitted = client.post(
            f"/api/student/assignments/{assignment_id}/submit",
            headers=headers,
            json={"work_url": f"https://github.com/example/{login}"},
        )
        assert submitted.status_code == 200

    for index in range(4):
        enroll_and_submit(f"balance_student_{index}")

    before = client.get(
        f"/api/assignments/{assignment_id}/reviewers", headers=methodist
    ).json()
    assert len(before) == 1
    assert before[0]["total"] == 4
    added = client.post(
        f"/api/assignments/{assignment_id}/reviewers",
        headers=methodist,
        json={"user_id": second["id"]},
    )
    assert added.status_code == 201
    after = client.get(
        f"/api/assignments/{assignment_id}/reviewers", headers=methodist
    ).json()
    totals = sorted(item["total"] for item in after)
    assert totals == [2, 2]
    removed = client.delete(
        f"/api/assignments/{assignment_id}/reviewers/{added.json()['id']}",
        headers=methodist,
    )
    assert removed.status_code == 204
    back = client.get(
        f"/api/assignments/{assignment_id}/reviewers", headers=methodist
    ).json()
    assert len(back) == 1
    assert back[0]["total"] == 4

    preview = client.post(
        f"/api/assignments/{assignment_id}/reviewers/import",
        headers=methodist,
        params={"confirm": "false"},
        files={
            "file": (
                "reviewers.xlsx",
                _xlsx_bytes("balance_two", "unknown_login"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert preview.status_code == 200
    body = preview.json()
    assert body["added"] == ["balance_two"]
    assert any("unknown_login" in item for item in body["errors"])
    assert body["applied"] is False
    still_one = client.get(
        f"/api/assignments/{assignment_id}/reviewers", headers=methodist
    ).json()
    assert len(still_one) == 1
    confirmed = client.post(
        f"/api/assignments/{assignment_id}/reviewers/import",
        headers=methodist,
        params={"confirm": "true"},
        files={
            "file": (
                "reviewers.xlsx",
                _xlsx_bytes("balance_two", "unknown_login"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert confirmed.json()["applied"] is True
    imported = client.get(
        f"/api/assignments/{assignment_id}/reviewers", headers=methodist
    ).json()
    assert sorted(item["total"] for item in imported) == [2, 2]

    course_xlsx = client.get(
        f"/api/courses/{course_id}/export.xlsx", headers=methodist
    )
    assert course_xlsx.status_code == 200
    course_book = load_workbook(BytesIO(course_xlsx.content))
    assert course_book.sheetnames == ["Students", "Reviewers", "Applications"]
    homework_xlsx = client.get(
        f"/api/assignments/{assignment_id}/export.xlsx", headers=methodist
    )
    assert homework_xlsx.status_code == 200
    homework_book = load_workbook(BytesIO(homework_xlsx.content))
    assert homework_book.sheetnames == ["Submissions", "Reviewers", "Scores"]

    course_preview = client.post(
        f"/api/courses/{course_id}/reviewers/import",
        headers=methodist,
        params={"confirm": "false"},
        files={
            "file": (
                "course.xlsx",
                _xlsx_bytes("ghost_reviewer"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert course_preview.status_code == 200
    assert course_preview.json()["added"] == []
    assert any("ghost_reviewer" in item for item in course_preview.json()["errors"])

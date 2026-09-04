from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def workbook_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write_sheet(
    workbook: Workbook,
    title: str,
    headers: list[str],
    rows: list[list[object]],
    *,
    first: bool = False,
) -> None:
    sheet: Worksheet = workbook.active if first else workbook.create_sheet(title)
    if first:
        sheet.title = title
    sheet.append(headers)
    for row in rows:
        sheet.append(row)


def export_course_workbook(
    *,
    students: list[list[object]],
    reviewers: list[list[object]],
    applications: list[list[object]],
) -> bytes:
    workbook = Workbook()
    _write_sheet(
        workbook,
        "Students",
        ["login", "first_name", "last_name", "telegram", "status"],
        students,
        first=True,
    )
    _write_sheet(
        workbook,
        "Reviewers",
        ["login", "first_name", "last_name", "telegram"],
        reviewers,
    )
    _write_sheet(
        workbook,
        "Applications",
        [
            "login",
            "first_name",
            "last_name",
            "telegram",
            "status",
            "created_at",
            "decided_at",
        ],
        applications,
    )
    return workbook_bytes(workbook)


def export_assignment_workbook(
    *,
    submissions: list[list[object]],
    reviewers: list[list[object]],
    scores: list[list[object]],
) -> bytes:
    workbook = Workbook()
    _write_sheet(
        workbook,
        "Submissions",
        [
            "student_name",
            "student_login",
            "work_url",
            "status",
            "reviewer",
            "reviewer_login",
            "score",
            "summary",
        ],
        submissions,
        first=True,
    )
    _write_sheet(
        workbook,
        "Reviewers",
        ["login", "name", "telegram", "checked", "total"],
        reviewers,
    )
    _write_sheet(
        workbook,
        "Scores",
        [
            "student_name",
            "student_login",
            "criterion",
            "score",
            "max_score",
            "comment",
            "total_score",
        ],
        scores,
    )
    return workbook_bytes(workbook)


def parse_logins(content: bytes) -> tuple[list[tuple[int, str]], list[str]]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception:
        return [], ["Не удалось прочитать файл. Нужен XLSX с колонкой login."]
    sheet = workbook.active
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1), [])
    headers = [
        str(cell.value).strip().lower() if cell.value is not None else ""
        for cell in header_cells
    ]
    if "login" not in headers:
        return [], ["В таблице нет обязательной колонки login."]
    login_index = headers.index("login")
    rows: list[tuple[int, str]] = []
    errors: list[str] = []
    seen: set[str] = set()
    for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if login_index >= len(row):
            continue
        raw = row[login_index].value
        if raw is None or str(raw).strip() == "":
            continue
        login = str(raw).strip()
        key = login.lower()
        if key in seen:
            errors.append(f"Строка {row_number}: повтор логина {login}")
            continue
        seen.add(key)
        rows.append((row_number, login))
    return rows, errors

import re
from pathlib import Path

import openpyxl
import requests

from .contracts import ExcelAudit, LinkInfo, SubmissionData
from .parsers import extract_task_text


class SubmissionParser:
    supported_extensions = frozenset({".py", ".ipynb", ".sql", ".sh", ".go", ".md", ".docx", ".pdf", ".xlsx", ".js", ".ts"})
    ignored_directories = frozenset({".git", "node_modules", ".venv", "venv", "__pycache__", "data", "datasets", "artifacts"})

    def build_from_local_repository(self, root: Path, submission_id: str, task_id: str) -> SubmissionData:
        files = [path for path in root.rglob("*") if path.is_file() and not any(part in self.ignored_directories for part in path.parts)]
        included = [path for path in files if path.suffix.lower() in self.supported_extensions]
        texts, tables, image_count = [], [], 0
        for path in included:
            parsed = self._parse_file(path)
            texts.append(f"## {path.relative_to(root)}\n{parsed['text']}")
            tables.extend(parsed["tables"])
            image_count += parsed["image_count"]
        raw_text = "\n\n".join(texts)
        return SubmissionData(
            submission_id=submission_id,
            task_id=task_id,
            file_type="github",
            file_tree=[str(path.relative_to(root)) for path in included],
            raw_text=raw_text,
            tables=tables,
            resolved_links=self._resolve_links(raw_text),
            image_count=image_count,
        )

    def parse_file(self, source: Path, submission_id: str, task_id: str) -> SubmissionData:
        parsed = self._parse_file(source)
        audit = self._audit_workbook(source) if source.suffix.lower() == ".xlsx" else None
        return SubmissionData(
            submission_id=submission_id,
            task_id=task_id,
            file_type=source.suffix.lower().lstrip("."),
            file_tree=[source.name],
            raw_text=parsed["text"],
            tables=parsed["tables"],
            excel_audit=audit,
            resolved_links=self._resolve_links(parsed["text"]),
            image_count=parsed["image_count"],
        )

    def _parse_file(self, source: Path) -> dict:
        suffix = source.suffix.lower()
        if suffix in {".pdf", ".docx", ".xlsx", ".md"}:
            text = extract_task_text(source)
            return {"text": text, "tables": [], "image_count": 0}
        if suffix == ".ipynb":
            return {"text": source.read_text(encoding="utf-8", errors="replace"), "tables": [], "image_count": 0}
        return {"text": source.read_text(encoding="utf-8", errors="replace"), "tables": [], "image_count": 0}

    def _audit_workbook(self, source: Path) -> ExcelAudit:
        workbook = openpyxl.load_workbook(source, data_only=False)
        formulas = hardcoded = rows = 0
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                if any(cell.value is not None for cell in row):
                    rows += 1
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas += 1
                    elif cell.value is not None:
                        hardcoded += 1
        return ExcelAudit(sheet_names=workbook.sheetnames, total_rows=rows, has_formulas=bool(formulas), hardcoded_count=hardcoded, formula_count=formulas)

    def _resolve_links(self, text: str) -> list[LinkInfo]:
        links = []
        for url in dict.fromkeys(re.findall(r"https?://[^\s<>\"]+", text)):
            try:
                response = requests.get(url, timeout=10, allow_redirects=True)
                links.append(LinkInfo(url=url, status_code=response.status_code, is_accessible=response.ok, content_summary=response.text[:500], is_google_doc="google" in url))
            except requests.RequestException as exc:
                links.append(LinkInfo(url=url, status_code=0, is_accessible=False, content_summary=str(exc), is_google_doc="google" in url))
        return links

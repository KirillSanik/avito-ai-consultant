"""FastAPI-приложение: единый эндпоинт ``POST /review`` (ТЗ §4).

``lifespan`` собирает сервисы из фабрик ``common.clients`` и кладёт
``Pipeline`` в ``app.state``. Эндпоинт принимает ``repo_url`` (Form) и
``task_file`` (UploadFile: PDF/DOCX/XLSX), пишет файл во временный каталог
(синхронная запись — в отдельном потоке) и возвращает ``ReviewResponse``
из ``common.models``.

Кодирования ошибок (ТЗ §4.2):
- недоступный репозиторий / неподдерживаемый файл — 422 с понятным сообщением
  (без токена и temp-пути — маскирует ``core.repo_clone``);
- сбой LLM (детектор/ревьюер/разбор условия) — 502;
- прочее — 500 (стандартное поведение FastAPI).
"""

from __future__ import annotations

import asyncio
import csv
import io
import logging
import shutil
import subprocess
import tempfile
import uuid
from contextlib import asynccontextmanager
from pathlib import Path
from urllib.parse import urlparse
from urllib.request import urlopen

from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from ai_detector.service import AIDetectionService
from common.clients import get_llm_client, get_openrouter_client
from common.db.models import (
    Course,
    CourseEnrollment,
    CourseReviewer,
    Evaluation,
    HomeworkProgress,
    Submission,
    SubmissionStatus,
    Task,
    User,
    UserRole,
)
from common.db.session import get_session, init_db, open_session
from common.models import Criterion as RubricCriterion
from common.models import TaskCriteria, TaskRubric
from common.parsers import SUPPORTED_TASK_EXTENSIONS, extract_task_text, parse_task_rubric
from common.settings import get_settings
from core.pipeline import Pipeline
from homework_reviewer.evaluator.grading_engine import GradingEngine
from homework_reviewer.reports.pdf_generator import generate_review_pdf

logger = logging.getLogger(__name__)

#: Префикс temp-каталога загруженных условий (тот же стиль, что у ``core.repo_clone.TEMP_DIR_PREFIX``).
TASK_FILE_PREFIX = "avito-review-task-"


def _error_response(status_code: int, detail: str) -> JSONResponse:
    return JSONResponse(status_code=status_code, content={"detail": detail})


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Старт: сборка Pipeline из фабрик клиентов; стоп: сброс состояния."""
    settings = get_settings()
    detector_service = AIDetectionService(get_llm_client(settings), settings)
    reviewer_service = GradingEngine(get_openrouter_client(settings), settings)
    app.state.pipeline = Pipeline(
        detector=detector_service,
        reviewer=reviewer_service,
        settings=settings,
    )
    storage_root = await asyncio.to_thread(lambda: Path(settings.storage_dir).resolve())
    for directory in (storage_root / "tasks", storage_root / "submissions", storage_root / "reports"):
        await asyncio.to_thread(directory.mkdir, parents=True, exist_ok=True)
    init_db(settings.database_url)
    _seed_demo_users()
    _seed_demo_content()
    _seed_demo_enrollments()
    _seed_homework_progress()
    app.state.storage_root = storage_root
    logger.info("Pipeline собран, сервис готов к обработке POST /review")
    yield
    app.state.pipeline = None
    logger.info("Pipeline освобождён")


app = FastAPI(title="Avito AI Consultant", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://localhost:3001",
        "http://127.0.0.1:3000",
        "http://127.0.0.1:3001",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class AuthRegisterPayload(BaseModel):
    first_name: str
    last_name: str
    login: str | None = None
    username: str | None = None
    password: str
    telegram: str = ""
    role: UserRole


class AuthLoginPayload(BaseModel):
    login: str
    password: str


class CoursePayload(BaseModel):
    title: str
    year: int = 2026
    cohort: str = ""
    stream: int = 1
    active: bool = True
    cover_color: str = "#3B6EF5"
    students_count: int = 0
    description: str = ""
    capacity: int = 30


class CourseDescriptionPayload(BaseModel):
    description: str = ""


class AssignmentGuidePayload(BaseModel):
    reviewer_guide: str = ""
    description: str | None = None


class ReviewerAssignmentPayload(BaseModel):
    reviewer_id: str | None = None
    login: str | None = None
    logins: list[str] = Field(default_factory=list)


def _auth_user_payload(user: User) -> dict[str, str]:
    return {
        "id": user.id,
        "login": user.id,
        "first_name": user.first_name or user.name.split(" ", 1)[0],
        "last_name": user.last_name,
        "telegram": user.telegram,
        "role": user.role.value,
    }


def _session_token(user: User) -> str:
    """Return a lightweight, local-only token for the frontend session."""
    return f"local-session:{user.id}"


def _seed_demo_users() -> None:
    """Keep the documented local demo accounts available on every fresh DB."""
    session = open_session()
    try:
        for login, role, first_name, password in (
            ("reviewer", UserRole.REVIEWER, "Demo reviewer", "reviewer"),
            ("methodist", UserRole.METHODIST, "Demo methodist", "methodist"),
            *[(f"student{number}", UserRole.STUDENT, f"Student {number}", "password") for number in range(1, 6)],
        ):
            user = session.get(User, login)
            if user is None:
                session.add(
                    User(
                        id=login,
                        role=role,
                        name=first_name,
                        first_name=first_name,
                        password=password,
                    )
                )
            elif not user.password:
                # Upgrade a demo user created by the pre-auth schema without
                # replacing a password that an operator has explicitly set.
                user.password = password
                user.first_name = user.first_name or first_name
                user.name = user.name or first_name
        session.commit()
    finally:
        session.close()


def _seed_demo_content() -> None:
    """Populate a fresh local install with cards the dashboard can display."""
    session = open_session()
    try:
        samples = (
            ("python-analysis", "Python для анализа данных", "Осенний поток", "#3B6EF5"),
            ("product-analytics", "Продуктовая аналитика", "Весенний поток", "#059669"),
        )
        for course_id, title, cohort, color in samples:
            course = session.get(Course, course_id)
            if course is None:
                course = Course(
                    id=course_id,
                    title=title,
                    cohort=cohort,
                    cover_color=color,
                    students_count=24,
                    description="Учебный курс с практическими домашними заданиями.",
                )
                session.add(course)
            task_id = f"{course_id}-intro"
            if session.get(Task, task_id) is None:
                rubric = TaskRubric(
                    task_id=task_id,
                    title="Первое практическое задание",
                    description="Подготовьте и сдайте решение по материалам занятия.",
                    full_instructions="Подготовьте решение и приложите ссылку на репозиторий или файл.",
                    criteria=[
                        RubricCriterion(
                            name="Полнота решения",
                            description="Решение соответствует условию.",
                            max_points=100,
                        )
                    ],
                    total_points=100,
                ).model_dump(mode="json")
                rubric.update({"deadline": "2026-09-20T23:59:00", "source_url": ""})
                session.add(Task(id=task_id, course_id=course_id, title=rubric["title"], rubric_json=rubric))
        session.commit()
    finally:
        session.close()


def _seed_demo_enrollments() -> None:
    """Enroll the five local demo students in every current course."""
    session = open_session()
    try:
        students = session.query(User).filter(User.role == UserRole.STUDENT).all()
        for course in session.query(Course).all():
            for student in students:
                exists = (
                    session.query(CourseEnrollment)
                    .filter(CourseEnrollment.course_id == course.id, CourseEnrollment.student_id == student.id)
                    .first()
                )
                if exists is None:
                    session.add(CourseEnrollment(id=str(uuid.uuid4()), course_id=course.id, student_id=student.id))
        session.commit()
    finally:
        session.close()


def _seed_homework_progress() -> None:
    """Create a zero-progress record for each enrolled student and task."""
    session = open_session()
    try:
        for task in session.query(Task).all():
            student_ids = [
                enrollment.student_id
                for enrollment in session.query(CourseEnrollment)
                .filter(CourseEnrollment.course_id == task.course_id)
                .all()
            ]
            for student_id in student_ids:
                exists = (
                    session.query(HomeworkProgress)
                    .filter(HomeworkProgress.task_id == task.id, HomeworkProgress.student_id == student_id)
                    .first()
                )
                if exists is None:
                    session.add(HomeworkProgress(id=str(uuid.uuid4()), task_id=task.id, student_id=student_id))
        session.commit()
    finally:
        session.close()


def _course_payload(course: Course, assignments_count: int, students_count: int | None = None) -> dict:
    return {
        "id": course.id,
        "title": course.title,
        "year": course.year,
        "cohort": course.cohort,
        "stream": course.stream,
        "active": course.active,
        "cover_color": course.cover_color,
        "students_count": students_count if students_count is not None else course.students_count,
        "assignments_count": assignments_count,
        "description": course.description,
        "capacity": course.capacity,
    }


def _homework_payload(session: Session, task: Task, number: int) -> dict:
    rubric = task.rubric_json
    return {
        "id": task.id,
        "course_id": task.course_id,
        "title": task.title,
        "description": rubric.get("description", ""),
        "number": number,
        "deadline": rubric.get("deadline", ""),
        "task_url": rubric.get("source_url", ""),
        "criteria": [
            {
                "title": item.get("name", item.get("title", "Критерий")),
                "description": item.get("description", ""),
                "max_score": item.get("max_points", item.get("max_score", 0)),
            }
            for item in rubric.get("criteria", [])
        ],
        "total": session.query(HomeworkProgress).filter(HomeworkProgress.task_id == task.id).count(),
        "reviewed": session.query(HomeworkProgress)
        .filter(HomeworkProgress.task_id == task.id, HomeworkProgress.submitted.is_(True))
        .count(),
    }


def _reviewer_payload(user: User, assignment: CourseReviewer | None = None) -> dict:
    return {
        "id": assignment.id if assignment else user.id,
        "user_id": user.id,
        "login": user.id,
        "name": f"{user.first_name} {user.last_name}".strip() or user.name or user.id,
        "first_name": user.first_name,
        "last_name": user.last_name,
        "telegram": user.telegram,
    }


def _course_reviewer(session: Session, course_id: str, reviewer_id: str) -> tuple[CourseReviewer, User]:
    user = session.get(User, reviewer_id)
    if user is None or user.role != UserRole.REVIEWER:
        raise HTTPException(422, "Указанный пользователь не является ревьюером")
    assignment = (
        session.query(CourseReviewer)
        .filter(CourseReviewer.course_id == course_id, CourseReviewer.reviewer_id == reviewer_id)
        .one_or_none()
    )
    if assignment is None:
        assignment = CourseReviewer(id=str(uuid.uuid4()), course_id=course_id, reviewer_id=reviewer_id)
        session.add(assignment)
    return assignment, user


def _optional_session_user(request: Request, session: Session) -> User | None:
    """Resolve the lightweight local session token without requiring auth for dashboard demos."""
    authorization = request.headers.get("Authorization", "")
    prefix = "Bearer local-session:"
    if not authorization.startswith(prefix):
        return None
    return session.get(User, authorization.removeprefix(prefix))


def _safe_id(value: str, field: str) -> str:
    if not value or Path(value).name != value or value in {".", ".."}:
        raise HTTPException(422, f"{field} должен быть непустым идентификатором без пути")
    return value


async def _store_upload(upload: UploadFile, directory: Path, entity_id: str) -> Path:
    suffix = Path(upload.filename or "").suffix.lower()
    if not suffix:
        raise HTTPException(422, "У загруженного файла должно быть расширение")
    path = directory / f"{entity_id}{suffix}"
    try:
        await asyncio.to_thread(path.write_bytes, await upload.read())
    except OSError as exc:
        raise HTTPException(500, "Не удалось сохранить загруженный файл") from exc
    return path


async def _download_task_file(url: str, directory: Path, entity_id: str) -> Path:
    parsed = urlparse(url)
    suffix = Path(parsed.path).suffix.lower()
    if parsed.scheme not in {"http", "https"} or suffix not in SUPPORTED_TASK_EXTENSIONS:
        raise HTTPException(422, f"url должен быть HTTP(S)-ссылкой на {', '.join(sorted(SUPPORTED_TASK_EXTENSIONS))}")

    def _download() -> bytes:
        with urlopen(url, timeout=20) as response:
            return response.read()

    try:
        data = await asyncio.to_thread(_download)
        path = directory / f"{entity_id}{suffix}"
        await asyncio.to_thread(path.write_bytes, data)
        return path
    except OSError as exc:
        raise HTTPException(422, "Не удалось скачать файл условия по url") from exc


async def _make_uploaded_solution_repo(source: Path, submission_id: str) -> tuple[Path, Path]:
    """Materialize one uploaded source file as a disposable Git repository."""
    temp_root = Path(tempfile.mkdtemp(prefix="avito-uploaded-submission-"))
    repo = temp_root / "repo"
    repo.mkdir()
    await asyncio.to_thread(shutil.copy2, source, repo / source.name)
    # The detector consumes text/code files. Preserve a textual companion for
    # office-format uploads so a DOCX/PDF/XLSX submission can use the same
    # repository pipeline as a source-code submission.
    if source.suffix.lower() in {".pdf", ".docx", ".xlsx"}:
        extracted = await asyncio.to_thread(extract_task_text, source)
        await asyncio.to_thread((repo / "submission_content.md").write_text, extracted, encoding="utf-8")

    def _git() -> None:
        for args in (
            ("git", "init"),
            ("git", "add", source.name),
            (
                "git", "-c", "user.name=Upload", "-c", "user.email=upload@local", "commit", "-m",
                f"submission {submission_id}",
            ),
        ):
            subprocess.run(args, cwd=repo, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

    try:
        await asyncio.to_thread(_git)
    except (OSError, subprocess.CalledProcessError) as exc:
        shutil.rmtree(temp_root, ignore_errors=True)
        raise HTTPException(500, "Не удалось подготовить загруженную сдачу к проверке") from exc
    return temp_root, repo


@app.get("/health")
async def health() -> dict[str, str]:
    """Liveness endpoint; does not call the database or an LLM."""
    return {"status": "ok"}


@app.post("/api/v1/auth/register", status_code=201)
async def auth_register(payload: AuthRegisterPayload, session: Session = Depends(get_session)) -> dict:
    """Create a local reviewer or methodist account for the frontend."""
    login = _safe_id((payload.login or payload.username or "").strip(), "login")
    if not payload.password:
        raise HTTPException(422, "Пароль не должен быть пустым")
    if payload.role not in {UserRole.REVIEWER, UserRole.METHODIST}:
        raise HTTPException(422, "Для регистрации доступны только reviewer и methodist")
    if session.get(User, login) is not None:
        raise HTTPException(409, "Пользователь с таким логином уже существует")
    user = User(
        id=login,
        role=payload.role,
        name=f"{payload.first_name.strip()} {payload.last_name.strip()}".strip(),
        first_name=payload.first_name.strip(),
        last_name=payload.last_name.strip(),
        telegram=payload.telegram.strip().removeprefix("@"),
        password=payload.password,
    )
    session.add(user)
    session.commit()
    return {"token": _session_token(user), "user": _auth_user_payload(user)}


@app.post("/api/v1/auth/login")
async def auth_login(payload: AuthLoginPayload, session: Session = Depends(get_session)) -> dict:
    """Authenticate a local demo account using its explicitly requested plaintext password."""
    login = _safe_id(payload.login.strip(), "login")
    user = session.get(User, login)
    if user is None or user.password != payload.password:
        raise HTTPException(401, "Неверный логин или пароль")
    return {"token": _session_token(user), "user": _auth_user_payload(user)}


@app.get("/api/v1/auth/me")
async def auth_me(request: Request, session: Session = Depends(get_session)) -> dict:
    authorization = request.headers.get("Authorization", "")
    prefix = "Bearer local-session:"
    if not authorization.startswith(prefix):
        raise HTTPException(401, "Требуется авторизация")
    user = session.get(User, authorization.removeprefix(prefix))
    if user is None:
        raise HTTPException(401, "Сессия недействительна")
    return _auth_user_payload(user)


@app.post("/api/v1/users/register", status_code=201)
async def register_user(
    user_id: str = Form(...), role: UserRole = Form(...), name: str = Form(...), session: Session = Depends(get_session)
) -> dict:
    user_id = _safe_id(user_id, "user_id")
    if session.get(User, user_id):
        raise HTTPException(409, "Пользователь с таким user_id уже существует")
    session.add(User(id=user_id, role=role, name=name))
    session.commit()
    return {"id": user_id, "role": role.value, "name": name}


@app.get("/api/v1/courses")
async def list_courses(session: Session = Depends(get_session)) -> list[dict]:
    courses = session.query(Course).order_by(Course.title).all()
    return [
        _course_payload(
            course,
            session.query(Task).filter(Task.course_id == course.id).count(),
            session.query(CourseEnrollment).filter(CourseEnrollment.course_id == course.id).count(),
        )
        for course in courses
    ]


@app.get("/api/v1/reviewers")
async def list_reviewers(session: Session = Depends(get_session)) -> list[dict]:
    reviewers = session.query(User).filter(User.role == UserRole.REVIEWER).order_by(User.id).all()
    return [_reviewer_payload(user) for user in reviewers]


@app.get("/api/v1/assignments/{homework_id}/reviewers")
async def list_assignment_reviewers(homework_id: str, session: Session = Depends(get_session)) -> list[dict]:
    """Return the parent course's reviewer team for the legacy assignment screen."""
    task = session.get(Task, _safe_id(homework_id, "homework_id"))
    if task is None:
        raise HTTPException(404, "Домашнее задание не найдено")
    assignments = session.query(CourseReviewer).filter(CourseReviewer.course_id == task.course_id).all()
    return [
        _reviewer_payload(user, assignment)
        for assignment in assignments
        if (user := session.get(User, assignment.reviewer_id)) is not None
    ]


@app.get("/api/v1/dashboard")
async def dashboard_summary(request: Request, session: Session = Depends(get_session)) -> dict:
    """Role-tolerant course and review summary used by the imported frontend."""
    user = _optional_session_user(request, session)
    courses_query = session.query(Course).filter(Course.active.is_(True))
    if user and user.role == UserRole.REVIEWER:
        course_ids = [
            item.course_id
            for item in session.query(CourseReviewer).filter(CourseReviewer.reviewer_id == user.id).all()
        ]
        courses_query = courses_query.filter(Course.id.in_(course_ids))
    elif user and user.role == UserRole.STUDENT:
        course_ids = [
            item.course_id
            for item in session.query(CourseEnrollment).filter(CourseEnrollment.student_id == user.id).all()
        ]
        courses_query = courses_query.filter(Course.id.in_(course_ids))
    courses = courses_query.order_by(Course.title).all()
    course_ids = [course.id for course in courses]
    tasks = session.query(Task).filter(Task.course_id.in_(course_ids)).all() if course_ids else []
    task_ids = [task.id for task in tasks]
    pending_reviews_count = (
        session.query(Submission)
        .filter(Submission.task_id.in_(task_ids), Submission.status == SubmissionStatus.PENDING)
        .count()
        if task_ids
        else 0
    )
    progress_total = (
        session.query(HomeworkProgress).filter(HomeworkProgress.task_id.in_(task_ids)).count()
        if task_ids
        else 0
    )
    progress_done = (
        session.query(HomeworkProgress)
        .filter(HomeworkProgress.task_id.in_(task_ids), HomeworkProgress.submitted.is_(True))
        .count()
        if task_ids
        else 0
    )
    course_payloads = [
        _course_payload(
            course,
            sum(task.course_id == course.id for task in tasks),
            session.query(CourseEnrollment).filter(CourseEnrollment.course_id == course.id).count(),
        )
        for course in courses
    ]
    return {
        "courses": course_payloads,
        "pending_reviews_count": pending_reviews_count,
        "active_homeworks_count": len(tasks),
        # Legacy dashboard keys consumed by older imported components.
        "total": progress_total,
        "reviewed": progress_done,
        "in_progress": max(progress_total - progress_done, 0),
        "reviewers": [],
        "clarifications": [],
    }


@app.get("/api/v1/courses/{course_id}/reviewers")
async def list_course_reviewers(course_id: str, session: Session = Depends(get_session)) -> list[dict]:
    course_id = _safe_id(course_id, "course_id")
    if session.get(Course, course_id) is None:
        raise HTTPException(404, "Курс не найден")
    assignments = session.query(CourseReviewer).filter(CourseReviewer.course_id == course_id).all()
    return [
        _reviewer_payload(session.get(User, assignment.reviewer_id), assignment)
        for assignment in assignments
        if session.get(User, assignment.reviewer_id) is not None
    ]


@app.post("/api/v1/courses/{course_id}/reviewers", status_code=201)
async def assign_course_reviewer(
    course_id: str, payload: ReviewerAssignmentPayload, session: Session = Depends(get_session)
) -> dict:
    course_id = _safe_id(course_id, "course_id")
    if session.get(Course, course_id) is None:
        raise HTTPException(404, "Курс не найден")
    reviewer_id = (payload.reviewer_id or payload.login or "").strip()
    if not reviewer_id:
        raise HTTPException(422, "Укажите reviewer_id или login")
    assignment, user = _course_reviewer(session, course_id, reviewer_id)
    session.commit()
    return _reviewer_payload(user, assignment)


@app.delete("/api/v1/courses/{course_id}/reviewers/{reviewer_id}", status_code=204)
async def unassign_course_reviewer(
    course_id: str, reviewer_id: str, session: Session = Depends(get_session)
) -> None:
    course_id, reviewer_id = _safe_id(course_id, "course_id"), _safe_id(reviewer_id, "reviewer_id")
    assignment = (
        session.query(CourseReviewer)
        .filter(CourseReviewer.course_id == course_id, CourseReviewer.reviewer_id == reviewer_id)
        .one_or_none()
    )
    if assignment is None:
        raise HTTPException(404, "Ревьюер не назначен на этот курс")
    session.delete(assignment)
    session.commit()


@app.post("/api/v1/courses/{course_id}/reviewers/import")
async def import_course_reviewers(request: Request, course_id: str, session: Session = Depends(get_session)) -> dict:
    """Assign existing reviewer accounts listed in JSON, CSV, or XLSX input."""
    course_id = _safe_id(course_id, "course_id")
    if session.get(Course, course_id) is None:
        raise HTTPException(404, "Курс не найден")
    content_type = request.headers.get("content-type", "")
    if content_type.startswith("application/json"):
        payload = ReviewerAssignmentPayload.model_validate(await request.json())
        logins = payload.logins
    else:
        form = await request.form()
        upload = form.get("file")
        if upload is None or not hasattr(upload, "read"):
            raise HTTPException(422, "Загрузите CSV или XLSX файл с колонкой login")
        raw = await upload.read()
        filename = str(getattr(upload, "filename", "")).lower()
        if filename.endswith(".csv"):
            logins = [row.get("login", "").strip() for row in csv.DictReader(io.StringIO(raw.decode("utf-8-sig")))]
        elif filename.endswith(".xlsx"):
            try:
                from openpyxl import load_workbook

                sheet = load_workbook(io.BytesIO(raw), read_only=True, data_only=True).active
                headers = [str(cell.value or "").strip().lower() for cell in next(sheet.iter_rows(max_row=1))]
                login_column = headers.index("login")
                logins = [str(row[login_column] or "").strip() for row in sheet.iter_rows(min_row=2, values_only=True)]
            except (ImportError, StopIteration, ValueError) as exc:
                raise HTTPException(422, "XLSX должен содержать колонку login") from exc
        else:
            raise HTTPException(422, "Поддерживаются только CSV и XLSX")
    assigned, missing = [], []
    for login in {login for login in logins if login}:
        user = session.get(User, login)
        if user is None or user.role != UserRole.REVIEWER:
            missing.append(login)
            continue
        assignment, _ = _course_reviewer(session, course_id, login)
        assigned.append(_reviewer_payload(user, assignment))
    session.commit()
    return {"assigned": assigned, "missing": missing}


@app.post("/api/v1/courses", status_code=201)
async def create_course(payload: CoursePayload, session: Session = Depends(get_session)) -> dict:
    course = Course(id=str(uuid.uuid4()), **payload.model_dump())
    session.add(course)
    for student in session.query(User).filter(User.role == UserRole.STUDENT).all():
        session.add(CourseEnrollment(id=str(uuid.uuid4()), course_id=course.id, student_id=student.id))
    session.commit()
    return _course_payload(course, 0, 5)


@app.patch("/api/v1/courses/{course_id}")
async def update_course_description(
    course_id: str, payload: CourseDescriptionPayload, session: Session = Depends(get_session)
) -> dict:
    course = session.get(Course, _safe_id(course_id, "course_id"))
    if course is None:
        raise HTTPException(404, "Курс не найден")
    course.description = payload.description.strip()
    session.commit()
    return _course_payload(
        course,
        session.query(Task).filter(Task.course_id == course.id).count(),
        session.query(CourseEnrollment).filter(CourseEnrollment.course_id == course.id).count(),
    )


@app.get("/api/v1/courses/{course_id}")
async def get_course(course_id: str, session: Session = Depends(get_session)) -> dict:
    course = session.get(Course, _safe_id(course_id, "course_id"))
    if course is None:
        raise HTTPException(404, "Курс не найден")
    return _course_payload(
        course,
        session.query(Task).filter(Task.course_id == course.id).count(),
        session.query(CourseEnrollment).filter(CourseEnrollment.course_id == course.id).count(),
    )


@app.get("/api/v1/courses/{course_id}/homeworks")
async def list_course_homeworks(course_id: str, session: Session = Depends(get_session)) -> list[dict]:
    course_id = _safe_id(course_id, "course_id")
    if session.get(Course, course_id) is None:
        raise HTTPException(404, "Курс не найден")
    tasks = session.query(Task).filter(Task.course_id == course_id).order_by(Task.title).all()
    return [_homework_payload(session, task, index) for index, task in enumerate(tasks, start=1)]


@app.get("/api/v1/assignments/{homework_id}")
@app.get("/api/v1/homeworks/{homework_id}")
async def get_homework_detail(homework_id: str, session: Session = Depends(get_session)) -> dict:
    """Compatibility detail route used by the Next.js homework card."""
    task = session.get(Task, _safe_id(homework_id, "homework_id"))
    if task is None:
        raise HTTPException(404, "Домашнее задание не найдено")
    rubric = task.rubric_json
    criteria = [
        {
            "title": item.get("name", item.get("title", "Критерий")),
            "description": item.get("description", ""),
            "max_score": item.get("max_points", item.get("max_score", 0)),
        }
        for item in rubric.get("criteria", [])
    ]
    return {
        "id": task.id,
        "course_id": task.course_id,
        "title": task.title,
        "description": rubric.get("description", ""),
        "deadline": rubric.get("deadline", ""),
        "task_url": rubric.get("source_url", ""),
        "file_path": task.file_path,
        "criteria": criteria,
        "reviewer_guide": rubric.get("full_instructions", ""),
        "submissions": [],
    }


@app.put("/api/v1/assignments/{homework_id}/criteria")
async def update_assignment_guide(
    homework_id: str, payload: AssignmentGuidePayload, session: Session = Depends(get_session)
) -> dict:
    """Save the reviewer manual and optional task description; rubric criteria remain backend-managed."""
    task = session.get(Task, _safe_id(homework_id, "homework_id"))
    if task is None:
        raise HTTPException(404, "Домашнее задание не найдено")
    rubric = dict(task.rubric_json)
    rubric["full_instructions"] = payload.reviewer_guide
    if payload.description is not None:
        rubric["description"] = payload.description
    task.rubric_json = rubric
    session.commit()
    return await get_homework_detail(homework_id, session)


@app.post("/api/v1/courses/{course_id}/homeworks", status_code=201)
async def create_course_homework(request: Request, course_id: str, session: Session = Depends(get_session)) -> dict:
    """Create a task from JSON or multipart data and retain its source in local storage."""
    course_id = _safe_id(course_id, "course_id")
    if session.get(Course, course_id) is None:
        raise HTTPException(404, "Курс не найден")
    content_type = request.headers.get("content-type", "")
    if content_type.startswith("application/json"):
        data = await request.json()
        upload = None
    else:
        form = await request.form()
        data = dict(form)
        upload = form.get("file")
    title = str(data.get("title", "")).strip()
    description = str(data.get("description") or data.get("instructions") or "").strip()
    source_url = str(data.get("url") or data.get("task_url") or "").strip()
    if not title or not description:
        raise HTTPException(422, "Укажите название и описание задания")
    if upload is not None and (not hasattr(upload, "filename") or not hasattr(upload, "read")):
        raise HTTPException(422, "file должен быть загруженным файлом")
    if bool(upload) == bool(source_url):
        raise HTTPException(422, "Укажите ровно один источник: file или url")
    task_id = str(uuid.uuid4())
    file_path: str | None = None
    if upload:
        saved = await _store_upload(upload, request.app.state.storage_root / "tasks", task_id)
        file_path = str(saved)
    rubric = TaskRubric(
        task_id=task_id,
        title=title,
        description=description,
        full_instructions=description,
        criteria=[
            RubricCriterion(
                name="Соответствие заданию",
                description="Решение соответствует инструкции.",
                max_points=100,
            )
        ],
        total_points=100,
    ).model_dump(mode="json")
    rubric.update({"deadline": str(data.get("deadline", "")), "source_url": source_url})
    task = Task(id=task_id, course_id=course_id, title=title, rubric_json=rubric, file_path=file_path)
    session.add(task)
    for enrollment in session.query(CourseEnrollment).filter(CourseEnrollment.course_id == course_id).all():
        session.add(HomeworkProgress(id=str(uuid.uuid4()), task_id=task_id, student_id=enrollment.student_id))
    session.commit()
    number = session.query(Task).filter(Task.course_id == course_id).count()
    return _homework_payload(session, task, number)


@app.post("/api/v1/tasks", status_code=201)
async def create_task(
    request: Request,
    course_id: str = Form(...),
    title: str = Form(...),
    file: UploadFile | None = File(None),
    url: str | None = Form(None),
    session: Session = Depends(get_session),
) -> dict:
    if bool(file) == bool(url):
        raise HTTPException(422, "Укажите ровно один источник условия: file или url")
    course_id = _safe_id(course_id, "course_id")
    task_id = str(uuid.uuid4())
    storage = request.app.state.storage_root / "tasks"
    source = (
        await _store_upload(file, storage, task_id)
        if file
        else await _download_task_file(url or "", storage, task_id)
    )
    if source.suffix.lower() not in SUPPORTED_TASK_EXTENSIONS:
        source.unlink(missing_ok=True)
        raise HTTPException(
            422,
            "Для API tasks поддерживаются только "
            f"{', '.join(sorted(SUPPORTED_TASK_EXTENSIONS))}",
        )
    try:
        text = await asyncio.to_thread(extract_task_text, source)
        rubric = await parse_task_rubric(
            text,
            task_id,
            client=request.app.state.pipeline._rubric_client(),
            settings=get_settings(),
        )
    except Exception:
        source.unlink(missing_ok=True)
        raise
    if session.get(Course, course_id) is None:
        session.add(Course(id=course_id, title=course_id))
    session.add(
        Task(
            id=task_id,
            course_id=course_id,
            title=title,
            rubric_json=rubric.model_dump(mode="json"),
            file_path=str(source),
        )
    )
    session.commit()
    return {"id": task_id, "course_id": course_id, "title": title, "rubric_json": rubric.model_dump(mode="json")}


@app.get("/api/v1/tasks/{task_id}")
async def get_task(task_id: str, session: Session = Depends(get_session)) -> dict:
    task_id = _safe_id(task_id, "task_id")
    task = session.get(Task, task_id)
    if task is None:
        raise HTTPException(404, "Задание не найдено")
    return {
        "id": task.id,
        "course_id": task.course_id,
        "title": task.title,
        "rubric_json": task.rubric_json,
        "file_path": task.file_path,
    }


@app.post("/api/v1/submissions", status_code=201)
async def create_submission(
    request: Request,
    task_id: str = Form(...),
    student_id: str = Form(...),
    repo_url: str | None = Form(None),
    file: UploadFile | None = File(None),
    session: Session = Depends(get_session),
) -> dict:
    if bool(repo_url) == bool(file):
        raise HTTPException(422, "Укажите ровно один источник сдачи: repo_url или file")
    task_id, student_id = _safe_id(task_id, "task_id"), _safe_id(student_id, "student_id")
    task, student = session.get(Task, task_id), session.get(User, student_id)
    if task is None or student is None:
        raise HTTPException(404, "Задание или студент не найдены")
    if student.role != UserRole.STUDENT:
        raise HTTPException(422, "student_id должен принадлежать пользователю с ролью student")
    submission_id = str(uuid.uuid4())
    saved_file: Path | None = None
    if file:
        saved_file = await _store_upload(file, request.app.state.storage_root / "submissions", submission_id)
    submission = Submission(
        id=submission_id,
        task_id=task_id,
        student_id=student_id,
        repo_url=repo_url,
        file_path=str(saved_file) if saved_file else None,
        status=SubmissionStatus.PROCESSING,
    )
    session.add(submission)
    session.commit()
    rubric = TaskRubric.model_validate(task.rubric_json)
    criteria = TaskCriteria(task_id=task_id, text=rubric.full_instructions, rubric=rubric)
    try:
        pipeline: Pipeline = request.app.state.pipeline
        if repo_url:
            review_result = await pipeline.run_preparsed(repo_url, criteria)
        else:
            temp_root, repo = await _make_uploaded_solution_repo(saved_file, submission_id)  # type: ignore[arg-type]
            try:
                review_result = await pipeline.run_from_path(f"upload://{submission_id}", criteria, repo)
            finally:
                shutil.rmtree(temp_root, ignore_errors=True)
        pdf_path = request.app.state.storage_root / "reports" / f"{submission_id}.pdf"
        await asyncio.to_thread(
            generate_review_pdf,
            review_result.evaluation,
            rubric,
            str(pdf_path),
            review_result.ai_assessment,
        )
        session.add(
            Evaluation(
                id=str(uuid.uuid4()),
                submission_id=submission_id,
                review_json=review_result.model_dump(mode="json"),
                pdf_path=str(pdf_path),
            )
        )
        submission.status = SubmissionStatus.COMPLETED
        session.commit()
    except Exception:
        submission.status = SubmissionStatus.FAILED
        session.commit()
        raise
    return {
        "submission_id": submission_id,
        "status": submission.status.value,
        "review_json": review_result.model_dump(mode="json"),
        "pdf_url": f"/api/v1/evaluations/{submission_id}/pdf",
    }


@app.get("/api/v1/submissions/{submission_id}")
async def get_submission(submission_id: str, session: Session = Depends(get_session)) -> dict:
    submission_id = _safe_id(submission_id, "submission_id")
    submission = session.get(Submission, submission_id)
    if submission is None:
        raise HTTPException(404, "Сдача не найдена")
    return {
        "id": submission.id,
        "task_id": submission.task_id,
        "student_id": submission.student_id,
        "repo_url": submission.repo_url,
        "file_path": submission.file_path,
        "status": submission.status.value,
    }


@app.get("/api/v1/evaluations")
async def get_evaluation(submission_id: str, session: Session = Depends(get_session)) -> dict:
    submission_id = _safe_id(submission_id, "submission_id")
    evaluation = session.query(Evaluation).filter(Evaluation.submission_id == submission_id).one_or_none()
    if evaluation is None:
        raise HTTPException(404, "Оценка не найдена")
    return {
        "submission_id": submission_id,
        "review_json": evaluation.review_json,
        "pdf_url": f"/api/v1/evaluations/{submission_id}/pdf",
    }


@app.get("/api/v1/evaluations/{submission_id}/pdf")
async def get_evaluation_pdf(submission_id: str, session: Session = Depends(get_session)) -> FileResponse:
    submission_id = _safe_id(submission_id, "submission_id")
    evaluation = session.query(Evaluation).filter(Evaluation.submission_id == submission_id).one_or_none()
    if evaluation is None or not await asyncio.to_thread(Path(evaluation.pdf_path).is_file):
        raise HTTPException(404, "PDF отчёт не найден")
    return FileResponse(evaluation.pdf_path, media_type="application/pdf", filename=f"{submission_id}.pdf")

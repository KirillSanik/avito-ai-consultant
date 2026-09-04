from pathlib import Path

from openpyxl import load_workbook

from src.models.submission import ExcelAudit


class XLSXParser:
    def parse(self, file_path: str) -> dict:
        source = Path(file_path)
        if not source.is_file():
            raise FileNotFoundError(f"Файл Excel не найден: {file_path}")
        formula_book = load_workbook(source, data_only=False)
        values_book = load_workbook(source, data_only=True)
        parts: list[str] = []
        tables: list[dict] = []
        links: list[str] = []
        hardcoded_count = 0
        formula_count = 0
        total_rows = 0
        image_count = 0
        for formula_sheet, values_sheet in zip(formula_book.worksheets, values_book.worksheets):
            image_count += len(formula_sheet._images)
            rows: list[list[str]] = []
            has_content = False
            for formula_row, value_row in zip(formula_sheet.iter_rows(), values_sheet.iter_rows()):
                row_values: list[str] = []
                row_has_content = False
                for formula_cell, value_cell in zip(formula_row, value_row):
                    formula_value = formula_cell.value
                    display_value = value_cell.value if isinstance(formula_value, str) and formula_value.startswith("=") else formula_value
                    if display_value is None and formula_value is not None:
                        display_value = formula_value
                    if formula_value is not None:
                        row_has_content = True
                        has_content = True
                        if isinstance(formula_value, str) and formula_value.startswith("="):
                            formula_count += 1
                        else:
                            hardcoded_count += 1
                    if formula_cell.hyperlink and formula_cell.hyperlink.target:
                        links.append(formula_cell.hyperlink.target)
                    row_values.append("" if display_value is None else str(display_value))
                if row_has_content:
                    total_rows += 1
                    rows.append(row_values)
            if has_content:
                tables.append({"sheet": formula_sheet.title, "rows": rows})
                parts.append(f"## Лист: {formula_sheet.title}\n{self._to_markdown(rows)}")
        audit = ExcelAudit(sheet_names=formula_book.sheetnames, total_rows=total_rows, has_formulas=formula_count > 0, hardcoded_count=hardcoded_count, formula_count=formula_count)
        formula_book.close()
        values_book.close()
        return {"raw_text": "\n\n".join(parts), "tables": tables, "links": links, "excel_audit": audit, "image_count": image_count}

    @staticmethod
    def _to_markdown(rows: list[list[str]]) -> str:
        if not rows:
            return ""
        width = max(len(row) for row in rows)
        normalized = [row + [""] * (width - len(row)) for row in rows]
        render = lambda row: "| " + " | ".join(value.replace("|", "\\|").replace("\n", " ") for value in row) + " |"
        return "\n".join([render(normalized[0]), render(["-" * 3] * width), *[render(row) for row in normalized[1:]]])
